

# Capturing the execl sheet path
import openpyxl

book = openpyxl.load_workbook("/home/chaitanya/Documents/PythonTesting/PythonSelenium/PythonDemo.xlsx")

# Monitoring the active sheet out of all the sheets present
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

# Writing value to the particular cell
(book.active.cell(row=2, column=2)).value = "Chaithanya"
book.active.cell(row=2, column=2, value="Sushma")

# Reading value from particular cell
print(book.active.cell(row=2, column=2).value)

print("row", book.active.max_row)
print("column", book.active.max_column)

# To get value using cell name
print(book.active['A5'].value)
print(book.active['A6'].value)

# Accessing the range of cells
print(book.active['A5':'C2'])

# Reading all the values present in the sheet
for i in range(1, book.active.max_row + 1):
    if book.active.cell(row=i, column=1).value == "testcase2":
        for j in range(1, book.active.max_column + 1):
            print(book.active.cell(row=i, column=j).value)
# Copying data from execle and storing it in the dictionary
Dict = {}

book = openpyxl.load_workbook("/home/chaitanya/Documents/PythonTesting/PythonSelenium/PythonDemo.xlsx")

for i in range(1, book.active.max_row + 1):  # to get rows
    if book.active.cell(row=i, column=1).value == "TestCase2":
        for j in range(1, book.active.max_column + 1):  # to get columns
            Dict[book.active.cell(row=1, column=j).value] = book.active.cell(row=i, column=j).value

print([Dict])
